import collections
import logging
import os
import re

from PyQt5 import QtCore, QtGui

import openpyxl

import tabula

import numpy as np

import pandas as pd


class RawDataError(Exception):
    """This class implements exceptions related with the contents of the data.
    """


class RawDataModel(QtCore.QAbstractTableModel):

    def __init__(self, *args, **kwargs):
        """Constructor.
        """

        super(RawDataModel, self).__init__(*args, **kwargs)

        self._rawdata = pd.DataFrame()

    def add_data(self, pdf_file, sort=False):
        """Add new data to the model.

        Args:
            data (pandas.DataFrame): the data
        """

        basename = os.path.basename(pdf_file)

        match = re.match(r'(\d{4}-\d{2}-\d{2}) .*(RT(1|2|3|1-2))_(\w+).PDF', basename)

        if match is None:
            raise IOError('Invalid filename')

        date, rt, _, gene = match.groups()

        pages = tabula.read_pdf(pdf_file, pages='all')

        data_frame = pd.DataFrame()

        # Loop over the table stored in each page of the pdf document
        for df in pages:
            data_frame = pd.concat([data_frame, df], ignore_index=True)

        # Drop unused columns
        for col in ['Inc', 'Type', 'Concentration', 'Standard', 'Status']:
            data_frame.drop(col, inplace=True, axis=1)

        n_samples = len(data_frame.index)

        # Clean up the Name column from leading "Standard" and "Control" strings
        for i in range(n_samples):
            data_frame['Name'].iloc[i] = ' '.join(data_frame['Name'].iloc[i].split()[1:]).strip()

        data_frame.insert(0, 'Date', [date]*n_samples)

        data_frame.insert(1, 'Gene', [gene]*n_samples)

        data_frame.insert(2, 'RT', [rt]*n_samples)

        data_frame.insert(6, 'File', [basename]*n_samples)

        data_frame['Date'] = pd.to_datetime(data_frame['Date'])
        data_frame['CP'] = data_frame['CP'].str.replace(',', '.').astype(np.float)

        self._rawdata = pd.concat([self._rawdata, data_frame])

        if sort:
            self.sort()

    def columnCount(self, parent=None):
        """Return the number of columns of the model for a given parent.

        Returns:
            int: the number of columns
        """

        return len(self._rawdata.columns)

    def data(self, index, role):
        """Get the data at a given index for a given role.

        Args:
            index (QtCore.QModelIndex): the index
            role (int): the role

        Returns:
            QtCore.QVariant: the data
        """

        if self._rawdata.empty:
            return QtCore.QVariant()

        if not index.isValid():
            return QtCore.QVariant()

        row = index.row()
        col = index.column()

        if role == QtCore.Qt.DisplayRole:
            return str(self._rawdata.iloc[row, col])
        elif role == QtCore.Qt.ForegroundRole:
            cp_value = self._rawdata['CP'].iloc[row]
            if np.isnan(cp_value):
                return QtGui.QBrush(QtCore.Qt.red)

    def export(self, workbook):
        """Export the raw data to an excel spreadsheet.

        Args:
            workbook (openpyxl.workbook.workbook.Workbook): the workbook
        """

        workbook.create_sheet('raw data')
        worksheet = workbook.get_sheet_by_name('raw data')

        for i, v in enumerate(self._rawdata.columns):
            worksheet.cell(row=1, column=i+2).value = v

        for i, v in enumerate(self._rawdata.index):
            worksheet.cell(row=i+2, column=1).value = v

        for i in range(len(self._rawdata.index)):
            for j in range(len(self._rawdata.columns)):
                worksheet.cell(row=i+2, column=j+2).value = self._rawdata.iloc[i, j]

    def headerData(self, col, orientation, role):
        """Returns the header data for a given row/column, orientation and role
        """

        if role == QtCore.Qt.DisplayRole:
            if orientation == QtCore.Qt.Horizontal:
                return self._rawdata.columns[col]
            else:
                return str(col+1)
        return None

    @property
    def rawdata(self):
        """Return the raw data.

        Returns:
            pandas.DataFrame: the raw data
        """

        return self._rawdata

    @rawdata.setter
    def rawdata(self, rawdata):
        """Setter for rawdata.

        Args:
            rawdata (pandas.DataFrame); the raw data
        """

        self._rawdata = rawdata

    def rowCount(self, parent=None):
        """Return the number of rows of the model for a given parent.

        Returns:
            int: the number of rows
        """

        return len(self._rawdata.index)

    @property
    def samples(self):
        """Return the samples names stored in the raw data.

        Returns:
            list of str: the sample names
        """

        samples = list(collections.OrderedDict.fromkeys(self._rawdata['Name']))

        return samples

    def sort(self):
        """Sort the raw data.
        """

        if self._rawdata.empty:
            return

        if 'Gene' not in self._rawdata.columns or 'Date' not in self._rawdata.columns:
            raise RawDataError('"Gene" or "Date" columns are missing from the raw data')

        self._rawdata.sort_values(by=['Gene', 'Date'], inplace=True, ascending=[True, True])
